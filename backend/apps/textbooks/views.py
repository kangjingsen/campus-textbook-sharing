from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import Q, F, Count, Sum, Case, When, IntegerField, Avg
from django.db import IntegrityError

from .models import Category, Textbook, TextbookVote, TextbookRating, TextbookComment, SharedResource, ResourceOrder
from .serializers import (
    CategorySerializer, CategoryFlatSerializer,
    TextbookListSerializer, TextbookDetailSerializer, TextbookCreateSerializer,
    TextbookCommentSerializer, SharedResourceSerializer, SharedResourceCreateSerializer,
    ResourceOrderSerializer, ResourceOrderCreateSerializer
)
from utils.permissions import IsOwnerOrAdmin, IsAdmin
from django.utils import timezone
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from decimal import Decimal, InvalidOperation
import csv
import io


CANCEL_REASONS = {item[0] for item in ResourceOrder.CANCEL_REASON_CHOICES}


class CategoryTreeView(generics.ListAPIView):
    """分类树形结构"""
    serializer_class = CategorySerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        return Category.objects.filter(parent__isnull=True, is_active=True)


class CategoryFlatListView(generics.ListAPIView):
    """分类扁平列表（用于下拉选择）"""
    queryset = Category.objects.filter(is_active=True)
    serializer_class = CategoryFlatSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = None


class CategoryManageView(generics.ListCreateAPIView):
    """管理员 - 分类管理"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdmin]
    pagination_class = None


class CategoryDetailManageView(generics.RetrieveUpdateDestroyAPIView):
    """管理员 - 分类详情/修改/删除"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdmin]


class TextbookListView(generics.ListAPIView):
    """教材列表（仅显示已审核通过的）"""
    serializer_class = TextbookListSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['transaction_type', 'condition', 'category', 'owner']
    search_fields = ['title', 'author', 'isbn', 'publisher']
    ordering_fields = ['price', 'created_at', 'view_count']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = Textbook.objects.filter(status__in=['approved', 'sold', 'rented', 'completed'])
        # 价格区间过滤
        min_price = self.request.query_params.get('min_price')
        max_price = self.request.query_params.get('max_price')
        if min_price:
            qs = qs.filter(price__gte=min_price)
        if max_price:
            qs = qs.filter(price__lte=max_price)
        return qs


class TextbookSearchView(APIView):
    """模糊搜索（按书名、作者、ISBN等）"""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        keyword = request.query_params.get('q', '').strip()
        if not keyword:
            return Response({'results': []})

        queryset = Textbook.objects.filter(
            status__in=['approved', 'sold', 'rented', 'completed']
        ).filter(
            Q(title__icontains=keyword) |
            Q(author__icontains=keyword) |
            Q(isbn__icontains=keyword) |
            Q(publisher__icontains=keyword) |
            Q(description__icontains=keyword)
        ).order_by('-view_count', '-created_at')[:50]

        serializer = TextbookListSerializer(queryset, many=True)
        return Response({'results': serializer.data, 'count': len(serializer.data)})


class TextbookCreateView(generics.CreateAPIView):
    """发布教材"""
    serializer_class = TextbookCreateSerializer
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        textbook = serializer.save()
        return Response({
            'message': '教材发布成功，等待审核',
            'textbook': TextbookDetailSerializer(textbook).data
        }, status=status.HTTP_201_CREATED)


class TextbookDetailView(generics.RetrieveAPIView):
    """教材详情（增加浏览次数，记录浏览历史）"""
    queryset = Textbook.objects.all()
    serializer_class = TextbookDetailSerializer
    permission_classes = [permissions.AllowAny]

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        # 增加浏览次数
        Textbook.objects.filter(pk=instance.pk).update(view_count=F('view_count') + 1)
        instance.refresh_from_db()

        # 记录浏览历史（用于推荐）
        if request.user.is_authenticated:
            from apps.recommendations.models import BrowsingHistory
            BrowsingHistory.objects.update_or_create(
                user=request.user,
                textbook=instance,
                defaults={}
            )

        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class TextbookUpdateView(generics.UpdateAPIView):
    """编辑教材（仅拥有者或管理员）"""
    queryset = Textbook.objects.all()
    serializer_class = TextbookCreateSerializer
    permission_classes = [IsOwnerOrAdmin]
    parser_classes = [MultiPartParser, FormParser]


class TextbookDeleteView(generics.DestroyAPIView):
    """删除教材（仅拥有者或管理员）"""
    queryset = Textbook.objects.all()
    permission_classes = [IsOwnerOrAdmin]


class MyTextbookListView(generics.ListAPIView):
    """我发布的教材"""
    serializer_class = TextbookListSerializer

    def get_queryset(self):
        queryset = Textbook.objects.filter(owner=self.request.user)
        keyword = (self.request.query_params.get('q') or '').strip()
        if keyword:
            queryset = queryset.filter(
                Q(title__icontains=keyword) |
                Q(author__icontains=keyword) |
                Q(isbn__icontains=keyword) |
                Q(publisher__icontains=keyword)
            )
        return queryset.order_by('-created_at')

    def list(self, request, *args, **kwargs):
        export_flag = str(request.query_params.get('export', '')).lower()
        if export_flag in ('1', 'true', 'yes'):
            export_view = TextbookBulkExportView()
            export_view.request = request
            return export_view.get(request)
        return super().list(request, *args, **kwargs)


class TextbookBulkExportView(APIView):
    """批量导出我的教材（csv/xlsx）"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not getattr(request.user, 'is_authenticated', False):
            return Response({'detail': '未登录'}, status=status.HTTP_401_UNAUTHORIZED)

        export_format = request.query_params.get('format', 'xlsx').lower()
        queryset = Textbook.objects.filter(owner=request.user).order_by('-created_at')
        keyword = (request.query_params.get('q') or '').strip()
        if keyword:
            queryset = queryset.filter(
                Q(title__icontains=keyword) |
                Q(author__icontains=keyword) |
                Q(isbn__icontains=keyword) |
                Q(publisher__icontains=keyword)
            )

        headers = [
            'id', 'title', 'author', 'isbn', 'publisher', 'edition',
            'condition', 'description', 'price', 'original_price',
            'transaction_type', 'rent_duration', 'category_id', 'category_name',
            'status', 'created_at'
        ]
        rows = []
        for tb in queryset:
            try:
                category_name = tb.category.name if tb.category else ''
            except ObjectDoesNotExist:
                category_name = ''

            rows.append([
                tb.id,
                tb.title,
                tb.author,
                tb.isbn,
                tb.publisher,
                tb.edition,
                tb.condition,
                tb.description,
                float(tb.price or 0),
                float(tb.original_price or 0) if tb.original_price is not None else '',
                tb.transaction_type,
                tb.rent_duration or '',
                tb.category_id or '',
                category_name,
                tb.status,
                tb.created_at.strftime('%Y-%m-%d %H:%M:%S') if tb.created_at else ''
            ])

        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="my_textbooks.csv"'
            response.write('\ufeff')
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
            return response

        try:
            from openpyxl import Workbook
        except ImportError:
            return Response({'error': '当前环境缺少 openpyxl，暂不支持 xlsx 导出，请改用 format=csv'}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active
        ws.title = 'textbooks'
        ws.append(headers)
        for row in rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="my_textbooks.xlsx"'
        return response


class TextbookBulkImportView(APIView):
    """批量导入教材（csv/xlsx）"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': '请上传文件 file'}, status=status.HTTP_400_BAD_REQUEST)

        filename = (upload.name or '').lower()
        try:
            if filename.endswith('.csv'):
                rows = self._read_csv(upload)
            elif filename.endswith('.xlsx'):
                rows = self._read_xlsx(upload)
            else:
                return Response({'error': '仅支持 .csv 或 .xlsx 文件'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': f'读取文件失败: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        errors = []
        for idx, row in enumerate(rows, start=2):
            try:
                payload = self._normalize_row(row)
                if not payload.get('title') or not payload.get('author'):
                    raise ValueError('title 和 author 为必填项')

                Textbook.objects.create(
                    title=payload['title'],
                    author=payload['author'],
                    isbn=payload.get('isbn', ''),
                    publisher=payload.get('publisher', ''),
                    edition=payload.get('edition', ''),
                    condition=payload.get('condition', 4),
                    description=payload.get('description', ''),
                    price=payload.get('price', Decimal('0')),
                    original_price=payload.get('original_price'),
                    transaction_type=payload.get('transaction_type', 'sell'),
                    rent_duration=payload.get('rent_duration'),
                    category=payload.get('category_obj'),
                    owner=request.user,
                    status='pending_review'
                )
                created_count += 1
            except Exception as exc:
                errors.append({'row': idx, 'error': str(exc)})

        return Response({
            'message': '导入完成',
            'created_count': created_count,
            'error_count': len(errors),
            'errors': errors[:50]
        })

    def _read_csv(self, upload):
        content = upload.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        return list(reader)

    def _read_xlsx(self, upload):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('当前环境缺少 openpyxl，暂不支持 xlsx 导入，请改用 csv')

        wb = load_workbook(upload, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        result = []
        for row in rows[1:]:
            result.append({headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))})
        return result

    def _normalize_row(self, row):
        payload = {
            'title': str(row.get('title', '') or '').strip(),
            'author': str(row.get('author', '') or '').strip(),
            'isbn': str(row.get('isbn', '') or '').strip(),
            'publisher': str(row.get('publisher', '') or '').strip(),
            'edition': str(row.get('edition', '') or '').strip(),
            'description': str(row.get('description', '') or '').strip(),
        }

        condition = row.get('condition', 4)
        try:
            payload['condition'] = int(condition) if condition not in (None, '') else 4
        except Exception:
            payload['condition'] = 4

        transaction_type = str(row.get('transaction_type', 'sell') or 'sell').strip()
        if transaction_type not in ('sell', 'rent', 'free'):
            transaction_type = 'sell'
        payload['transaction_type'] = transaction_type

        price = row.get('price', 0)
        try:
            payload['price'] = Decimal(str(price or 0))
        except (InvalidOperation, ValueError):
            payload['price'] = Decimal('0')

        original_price = row.get('original_price')
        if original_price in (None, ''):
            payload['original_price'] = None
        else:
            try:
                payload['original_price'] = Decimal(str(original_price))
            except (InvalidOperation, ValueError):
                payload['original_price'] = None

        rent_duration = row.get('rent_duration')
        if rent_duration in (None, ''):
            payload['rent_duration'] = None
        else:
            payload['rent_duration'] = int(rent_duration)

        category_obj = None
        category_id = row.get('category_id')
        category_name = str(row.get('category_name', '') or '').strip()
        if category_id not in (None, ''):
            category_obj = Category.objects.filter(id=int(category_id), is_active=True).first()
        elif category_name:
            category_obj = Category.objects.filter(name=category_name, is_active=True).first()
        payload['category_obj'] = category_obj

        if payload['transaction_type'] == 'free':
            payload['price'] = Decimal('0')

        return payload


# ============= 管理员删除 =============

class AdminTextbookDeleteView(generics.DestroyAPIView):
    """管理员删除任意教材"""
    queryset = Textbook.objects.all()
    permission_classes = [IsAdmin]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        title = instance.title
        self.perform_destroy(instance)
        return Response({'message': f'已删除教材「{title}」'}, status=status.HTTP_200_OK)


# ============= 点赞/点踩 =============

class TextbookVoteView(APIView):
    """教材点赞/点踩 — POST 切换投票"""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get(self, request, pk):
        """获取教材的投票统计和当前用户的投票状态"""
        likes = TextbookVote.objects.filter(textbook_id=pk, vote=1).count()
        dislikes = TextbookVote.objects.filter(textbook_id=pk, vote=-1).count()
        my_vote = 0
        if request.user.is_authenticated:
            v = TextbookVote.objects.filter(textbook_id=pk, user=request.user).first()
            if v:
                my_vote = v.vote
        return Response({'likes': likes, 'dislikes': dislikes, 'my_vote': my_vote})

    def post(self, request, pk):
        """投票：vote=1 (赞) 或 vote=-1 (踩)，再次相同则取消"""
        vote_val = request.data.get('vote')
        if vote_val not in (1, -1, '1', '-1'):
            return Response({'error': 'vote 必须为 1 或 -1'}, status=status.HTTP_400_BAD_REQUEST)
        vote_val = int(vote_val)

        existing = TextbookVote.objects.filter(textbook_id=pk, user=request.user).first()
        if existing:
            if existing.vote == vote_val:
                existing.delete()  # 取消投票
                action = 'cancelled'
            else:
                existing.vote = vote_val
                existing.save()
                action = 'changed'
        else:
            TextbookVote.objects.create(textbook_id=pk, user=request.user, vote=vote_val)
            action = 'voted'

        likes = TextbookVote.objects.filter(textbook_id=pk, vote=1).count()
        dislikes = TextbookVote.objects.filter(textbook_id=pk, vote=-1).count()
        my_vote = 0
        v = TextbookVote.objects.filter(textbook_id=pk, user=request.user).first()
        if v:
            my_vote = v.vote
        return Response({'action': action, 'likes': likes, 'dislikes': dislikes, 'my_vote': my_vote})


class TextbookRatingView(APIView):
    """教材星级评分（可修改、可取消）"""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get(self, request, pk):
        rating_qs = TextbookRating.objects.filter(textbook_id=pk)
        avg_value = rating_qs.aggregate(avg=Avg('score'))['avg']
        my_rating = 0
        if request.user.is_authenticated:
            mine = rating_qs.filter(user=request.user).first()
            if mine:
                my_rating = int(mine.score)
        return Response({
            'avg_score': round(float(avg_value), 2) if avg_value is not None else 0,
            'rating_count': rating_qs.count(),
            'my_rating': my_rating,
        })

    def post(self, request, pk):
        score = request.data.get('score')
        try:
            score = int(score)
        except (TypeError, ValueError):
            return Response({'error': 'score 必须为 0-5 的整数'}, status=status.HTTP_400_BAD_REQUEST)

        if score < 0 or score > 5:
            return Response({'error': 'score 必须为 0-5 的整数'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            textbook = Textbook.objects.get(pk=pk)
        except Textbook.DoesNotExist:
            return Response({'error': '教材不存在'}, status=status.HTTP_404_NOT_FOUND)

        if textbook.owner_id == request.user.id:
            return Response({'error': '不能给自己的教材评分'}, status=status.HTTP_400_BAD_REQUEST)

        existing = TextbookRating.objects.filter(textbook_id=pk, user=request.user).first()

        # score=0 或重复点击同分值 -> 取消评分
        if existing and (score == 0 or existing.score == score):
            existing.delete()
            action = 'cancelled'
            my_rating = 0
        elif existing:
            existing.score = score
            existing.save(update_fields=['score'])
            action = 'updated'
            my_rating = score
        else:
            if score == 0:
                return Response({'error': '当前尚未评分，无法取消'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                TextbookRating.objects.create(textbook_id=pk, user=request.user, score=score)
            except IntegrityError:
                return Response({'error': '评分提交失败，请重试'}, status=status.HTTP_400_BAD_REQUEST)
            action = 'created'
            my_rating = score

        qs = TextbookRating.objects.filter(textbook_id=pk)
        avg_value = qs.aggregate(avg=Avg('score'))['avg']
        count = qs.count()
        return Response({
            'message': '评分已更新' if action in ('updated', 'created') else '评分已取消',
            'action': action,
            'avg_score': round(float(avg_value), 2) if avg_value is not None else 0,
            'rating_count': count,
            'my_rating': my_rating,
        }, status=status.HTTP_200_OK)


# ============= 评论 =============

class TextbookCommentListView(generics.ListCreateAPIView):
    """教材评论列表 / 发表评论"""
    serializer_class = TextbookCommentSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        return TextbookComment.objects.filter(
            textbook_id=self.kwargs['pk'],
            parent__isnull=True  # 只返回顶级评论，回复嵌套在内
        )

    def perform_create(self, serializer):
        serializer.save(user=self.request.user, textbook_id=self.kwargs['pk'])


class TextbookCommentDeleteView(generics.DestroyAPIView):
    """删除评论（本人或管理员）"""
    queryset = TextbookComment.objects.all()
    permission_classes = [IsOwnerOrAdmin]

    def get_object(self):
        obj = super().get_object()
        # IsOwnerOrAdmin 需要 owner 属性
        obj.owner = obj.user
        return obj


# ============= 在线资料共享区 =============

class SharedResourceListView(generics.ListCreateAPIView):
    """资料列表 / 上传资料"""
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['resource_type', 'category', 'sale_type']
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'download_count']
    ordering = ['-created_at']
    parser_classes = [MultiPartParser, FormParser]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return SharedResourceCreateSerializer
        return SharedResourceSerializer

    def get_queryset(self):
        return SharedResource.objects.all()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        resource = serializer.save()
        return Response({
            'message': '资料上传成功',
            'resource': SharedResourceSerializer(resource).data
        }, status=status.HTTP_201_CREATED)


class MySharedResourceListView(generics.ListAPIView):
    """我上传的在线资料"""
    serializer_class = SharedResourceSerializer

    def get_queryset(self):
        queryset = SharedResource.objects.filter(uploader=self.request.user)
        keyword = (self.request.query_params.get('q') or '').strip()
        if keyword:
            queryset = queryset.filter(
                Q(title__icontains=keyword) |
                Q(description__icontains=keyword) |
                Q(category__name__icontains=keyword)
            )
        sale_type = self.request.query_params.get('sale_type')
        if sale_type:
            queryset = queryset.filter(sale_type=sale_type)
        resource_type = self.request.query_params.get('resource_type')
        if resource_type:
            queryset = queryset.filter(resource_type=resource_type)
        return queryset.order_by('-created_at')

    def list(self, request, *args, **kwargs):
        export_flag = str(request.query_params.get('export', '')).lower()
        if export_flag in ('1', 'true', 'yes'):
            export_view = MySharedResourceExportView()
            export_view.request = request
            return export_view.get(request)
        return super().list(request, *args, **kwargs)


class MySharedResourceExportView(APIView):
    """批量导出我的在线资料（csv/xlsx）"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        export_format = request.query_params.get('format', 'xlsx').lower()
        queryset = SharedResource.objects.filter(uploader=request.user).order_by('-created_at')

        keyword = (request.query_params.get('q') or '').strip()
        if keyword:
            queryset = queryset.filter(
                Q(title__icontains=keyword) |
                Q(description__icontains=keyword) |
                Q(category__name__icontains=keyword)
            )

        sale_type = request.query_params.get('sale_type')
        if sale_type:
            queryset = queryset.filter(sale_type=sale_type)

        resource_type = request.query_params.get('resource_type')
        if resource_type:
            queryset = queryset.filter(resource_type=resource_type)

        headers = [
            'id', 'title', 'description', 'resource_type', 'resource_type_display',
            'sale_type', 'sale_type_display', 'price', 'category_id', 'category_name',
            'download_count', 'file_size', 'created_at'
        ]
        rows = []
        for item in queryset:
            rows.append([
                item.id,
                item.title,
                item.description,
                item.resource_type,
                item.get_resource_type_display(),
                item.sale_type,
                item.get_sale_type_display(),
                float(item.price or 0),
                item.category_id or '',
                item.category.name if item.category else '',
                int(item.download_count or 0),
                int(item.file_size or 0),
                item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else ''
            ])

        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="my_resources.csv"'
            response.write('\ufeff')
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
            return response

        try:
            from openpyxl import Workbook
        except ImportError:
            return Response({'error': '当前环境缺少 openpyxl，暂不支持 xlsx 导出，请改用 format=csv'}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active
        ws.title = 'resources'
        ws.append(headers)
        for row in rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="my_resources.xlsx"'
        return response


class SharedResourceDetailView(generics.RetrieveDestroyAPIView):
    """资料详情 / 删除（本人或管理员）"""
    queryset = SharedResource.objects.all()
    serializer_class = SharedResourceSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [IsOwnerOrAdmin()]

    def get_object(self):
        obj = super().get_object()
        if self.request.method == 'DELETE':
            obj.owner = obj.uploader  # for IsOwnerOrAdmin
        return obj


class SharedResourceDownloadView(APIView):
    """记录下载次数"""
    permission_classes = [permissions.AllowAny]

    def post(self, request, pk):
        try:
            resource = SharedResource.objects.get(pk=pk)
        except SharedResource.DoesNotExist:
            return Response({'error': '资料不存在'}, status=status.HTTP_404_NOT_FOUND)

        can_download = resource.sale_type == 'free'
        if request.user.is_authenticated:
            if request.user.id == resource.uploader_id or request.user.role in ('admin', 'superadmin'):
                can_download = True
            elif ResourceOrder.objects.filter(resource=resource, buyer=request.user, status='completed').exists():
                can_download = True

        if not can_download:
            return Response({'error': '请先完成资料订单支付后下载'}, status=status.HTTP_403_FORBIDDEN)

        SharedResource.objects.filter(pk=pk).update(download_count=F('download_count') + 1)
        return Response({'message': 'ok', 'file': resource.file.url if resource.file else ''})


class ResourceOrderCreateView(APIView):
    """创建资料订单"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = ResourceOrderCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            resource = SharedResource.objects.get(pk=serializer.validated_data['resource_id'])
        except SharedResource.DoesNotExist:
            return Response({'error': '资料不存在'}, status=status.HTTP_404_NOT_FOUND)

        if resource.sale_type == 'free':
            return Response({'error': '免费资料无需下单'}, status=status.HTTP_400_BAD_REQUEST)

        if resource.uploader_id == request.user.id:
            return Response({'error': '不能购买自己的资料'}, status=status.HTTP_400_BAD_REQUEST)

        exists = ResourceOrder.objects.filter(
            resource=resource,
            buyer=request.user,
            status__in=['pending', 'confirmed', 'paid_pending', 'completed']
        ).exists()
        if exists:
            return Response({'error': '您已有该资料订单'}, status=status.HTTP_400_BAD_REQUEST)

        order = ResourceOrder.objects.create(
            resource=resource,
            buyer=request.user,
            seller=resource.uploader,
            price=resource.price,
            note=serializer.validated_data.get('note', '')
        )
        return Response({'message': '资料订单创建成功', 'order': ResourceOrderSerializer(order).data}, status=status.HTTP_201_CREATED)


class ResourceOrderListView(generics.ListAPIView):
    """资料订单列表"""
    serializer_class = ResourceOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        role = self.request.query_params.get('role', 'all')
        order_status = self.request.query_params.get('status')
        qs = ResourceOrder.objects.all()
        if role == 'buyer':
            qs = qs.filter(buyer=self.request.user)
        elif role == 'seller':
            qs = qs.filter(seller=self.request.user)
        else:
            qs = qs.filter(Q(buyer=self.request.user) | Q(seller=self.request.user))

        if order_status:
            qs = qs.filter(status=order_status)
        return qs


class ResourceOrderDetailView(generics.RetrieveAPIView):
    """资料订单详情"""
    serializer_class = ResourceOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return ResourceOrder.objects.filter(Q(buyer=self.request.user) | Q(seller=self.request.user))


class ResourceOrderConfirmView(APIView):
    """卖家确认资料订单并给出支付二维码"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def post(self, request, pk):
        qr = request.data.get('payment_qr', '').strip()
        qr_image = request.FILES.get('payment_qr_image')
        if not qr and not qr_image:
            return Response({'error': '请提供 payment_qr 或上传 payment_qr_image'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            order = ResourceOrder.objects.get(pk=pk, seller=request.user, status='pending')
        except ResourceOrder.DoesNotExist:
            return Response({'error': '订单不存在'}, status=status.HTTP_404_NOT_FOUND)

        order.status = 'confirmed'
        order.payment_qr = qr
        if qr_image:
            order.payment_qr_image = qr_image
        order.confirmed_at = timezone.now()
        update_fields = ['status', 'payment_qr', 'confirmed_at', 'updated_at']
        if qr_image:
            update_fields.append('payment_qr_image')
        order.save(update_fields=update_fields)
        return Response({'message': '订单已确认，支付二维码已发送'})


class ResourceOrderCompleteView(APIView):
    """买家确认支付完成"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        try:
            order = ResourceOrder.objects.get(pk=pk, buyer=request.user, status='confirmed')
        except ResourceOrder.DoesNotExist:
            return Response({'error': '订单不存在'}, status=status.HTTP_404_NOT_FOUND)

        proof = request.FILES.get('payment_proof')
        if not proof:
            return Response({'error': '请上传支付凭证'}, status=status.HTTP_400_BAD_REQUEST)

        order.status = 'paid_pending'
        order.payment_proof = proof
        order.paid_at = timezone.now()
        order.save(update_fields=['status', 'payment_proof', 'paid_at', 'updated_at'])
        return Response({'message': '支付凭证已提交，等待卖家确认'})


class ResourceOrderSellerCompleteView(APIView):
    """卖家确认收款并完成订单"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            order = ResourceOrder.objects.get(pk=pk, seller=request.user, status='paid_pending')
        except ResourceOrder.DoesNotExist:
            return Response({'error': '订单不存在'}, status=status.HTTP_404_NOT_FOUND)

        order.status = 'completed'
        order.completed_at = timezone.now()
        order.save(update_fields=['status', 'completed_at', 'updated_at'])
        return Response({'message': '已确认收款，订单完成，买家可下载'})


class ResourceOrderCancelView(APIView):
    """取消资料订单"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            order = ResourceOrder.objects.get(pk=pk, status__in=['pending', 'confirmed'])
            if order.buyer_id != request.user.id and order.seller_id != request.user.id:
                raise ResourceOrder.DoesNotExist
        except ResourceOrder.DoesNotExist:
            return Response({'error': '订单不存在或无法取消'}, status=status.HTTP_404_NOT_FOUND)

        reason = str(request.data.get('reason', '') or '').strip()
        if reason not in CANCEL_REASONS:
            return Response({'error': '取消原因无效，请选择预设原因'}, status=status.HTTP_400_BAD_REQUEST)

        cancel_by_role = 'buyer' if order.buyer_id == request.user.id else 'seller'
        order.status = 'cancelled'
        order.cancel_reason = reason
        order.cancel_by_role = cancel_by_role
        order.save(update_fields=['status', 'cancel_reason', 'cancel_by_role', 'updated_at'])
        return Response({'message': '已取消'})
